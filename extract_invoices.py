import os
import re
import pandas as pd
from PyPDF2 import PdfReader
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Input/output paths
input_folder = os.path.join("..", "Input")
output_file = os.path.join("..", "Output", "extracted_data.xlsx")

invoice_data = []

# Keywords
date_keywords = ["invoice date", "date of issue", "issue date"]
total_keywords = ["grand total", "total amount", "amount payable", "total", "total payable"]
invoice_keywords = ["invoice number", "invoice no", "tax invoice number", "inv no", "invoice #"]
seller_keywords = ["sold by", "seller", "supplied by", "sold to"]
order_keywords = ["order number", "order no", "order id"]
gst_regex = r"\d{2}[A-Z]{5}\d{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}"

# Date parser
def parse_and_format_date(date_string):
    date_formats = [
        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
        "%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d",
        "%B %d, %Y", "%b %d, %Y"
    ]
    for fmt in date_formats:
        try:
            date_obj = datetime.strptime(date_string, fmt)
            return date_obj.strftime("%d/%m/%Y")
        except ValueError:
            continue
    return date_string

# === PDF Processing ===
for filename in os.listdir(input_folder):
    if filename.endswith(".pdf"):
        filepath = os.path.join(input_folder, filename)
        reader = PdfReader(filepath)
        text = "".join(page.extract_text() for page in reader.pages if page.extract_text())

        invoice_number = ""
        invoice_date = ""
        total_amount = ""
        seller = ""
        order_number = ""
        gstin = ""

        lines = text.split("\n")
        for idx, line in enumerate(lines):
            line_clean = line.strip()
            line_lower = line_clean.lower()

            if not invoice_number:
                for key in invoice_keywords:
                    if key in line_lower:
                        match = re.search(r"(invoice\s*(number|no|#)[\s:#\-]*)([A-Z0-9\-\/]+)", line_clean, re.IGNORECASE)
                        if match:
                            invoice_number = match.group(3).strip()
                            break

            if not invoice_date:
                for key in date_keywords:
                    if key in line_lower:
                        match = re.search(
                            r"(\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4})|"
                            r"(\d{4}[\/\-.]\d{1,2}[\/\-.]\d{1,2})|"
                            r"([A-Za-z]+\s+\d{1,2},\s*\d{4})",
                            line_clean
                        )
                        if match:
                            raw_date = match.group().strip()
                            invoice_date = parse_and_format_date(raw_date)
                            break

            if not total_amount:
                for key in total_keywords:
                    if key in line_lower:
                        amount_match = re.search(r"₹\s?[\d,]+\.?\d*", line_clean)
                        if not amount_match:
                            amount_match = re.search(r"[\d,]+\.\d{2}", line_clean)
                        if amount_match:
                            total_amount = amount_match.group().replace("₹", "").strip()
                            break

            if not seller:
                for key in seller_keywords:
                    if key in line_lower:
                        seller = line_clean
                        break

            if not order_number:
                for key in order_keywords:
                    if key in line_lower:
                        match = re.search(r"[:\s\-#]*([A-Z0-9\-\/]{5,})", line_clean)
                        if match:
                            order_number = match.group(1).strip()
                            break

            if not gstin:
                gst_match = re.search(gst_regex, line_clean)
                if gst_match:
                    gstin = gst_match.group().strip()

        # Final clean conversion of amount
        try:
            clean_amount = float(total_amount.replace(",", "")) if total_amount else None
        except:
            clean_amount = None

        invoice_data.append({
            "File Name": filename,
            "Invoice Number": invoice_number if invoice_number else "Not Found",
            "Invoice Date": invoice_date if invoice_date else "Not Found",
            "Order Number": order_number if order_number else "Not Found",
            "GSTIN": gstin if gstin else "Not Found",
            "Total Amount (INR)": clean_amount if clean_amount is not None else "Not Found",
            "Seller": seller if seller else "Not Found"
        })

# === Save to Excel ===
if os.path.exists(output_file):
    try:
        os.remove(output_file)
    except PermissionError:
        print("❗ Close the Excel file before running again.")
        exit()

df = pd.DataFrame(invoice_data)
df = df[df["Invoice Number"] != "Not Found"]  # Remove invalid rows
df.to_excel(output_file, index=False)

# === Format Excel ===
wb = load_workbook(output_file)
ws = wb.active

# Format styles
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = 25
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if row == 1:
            cell.font = Font(bold=True)

# Freeze header
ws.freeze_panes = ws["A2"]
wb.save(output_file)

print(f"✅ Data extracted and formatted successfully to: {output_file}")
